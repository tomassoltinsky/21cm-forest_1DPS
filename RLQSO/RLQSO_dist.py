import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib import gridspec
from matplotlib.ticker import AutoMinorLocator
import openpyxl

def excel_column(data_name,n_col):

    wb = openpyxl.load_workbook(data_name)
    sheet = wb.active
    N_rows = sheet.max_row-1

    col = np.empty(N_rows)

    for i_row in range(N_rows):
        col[i_row] = sheet.cell(row=i_row+2,column=n_col).value

    return col

z = excel_column('radio_sources.xlsx',2)
S147 = excel_column('radio_sources.xlsx',6)
S147_limit = excel_column('radio_sources.xlsx',7)
DEC = excel_column('radio_sources.xlsx',17)

S147_limit[5] = 0
S147_limit[6] = 0

print(z)
print(S147)
print(S147_limit)

fsize = 16
fsize_leg = 14



fig = plt.figure(figsize=(5.,5.))
gs = gridspec.GridSpec(1,1)

ax = plt.subplot(gs[0,0])

ind_lim = np.where(S147_limit>0)[0]
ind_nolim = np.where(S147_limit<1)[0]

ax.errorbar(z[ind_lim], S147[ind_lim], yerr=0.2*S147[ind_lim], uplims=True, fmt='o', markersize=7.5, color='royalblue', label='Upper limit')
ax.errorbar(z[ind_nolim], S147[ind_nolim], uplims=False, fmt='o', markersize=7.5, color='royalblue', label='Detection')
ax.set_xlim(5.4,7.2)
ax.set_ylim(0.1,200)
ax.set_yscale('log')
ax.set_xlabel(r'$z$',fontsize=fsize)
ax.set_ylabel(r'$S_{147}\,\rm [mJy]$',fontsize=fsize)
ax.xaxis.set_minor_locator(AutoMinorLocator())
ax.tick_params(axis='both',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

plt.tight_layout()
plt.savefig('../plots/RLQSO_Svsz_log.png')
plt.show()
plt.close()



Smin_crit = 1.

ind_S147       = np.where(S147>=Smin_crit)[0]
ind_S147_wolim = np.where((S147>=Smin_crit) & (S147_limit<1))[0]

z_S147 = z[ind_S147]
z_S147_wolim = z[ind_S147_wolim]

print('Number of z>=5.5 RLQSO:    %d' % len(z))

ind_north = np.where(DEC>0)[0]
z_north = z[ind_north]
ind_north = np.where((DEC>0) & (S147_limit==0))[0]
S147_north = S147[ind_north]
print('Number of RLQSO in N hemi: %d' % len(z_north))

bins = np.arange(5.3,7.2,0.2)
fig = plt.figure(figsize=(5.,5.))
gs = gridspec.GridSpec(1,1)

ax = plt.subplot(gs[0,0])

ax.hist(z,bins=bins,density=False,histtype='step',linewidth=2,color='royalblue',label='All RLQSO')
ax.hist(z_north,bins=bins,density=False,histtype='step',linestyle=':',linewidth=2,color='fuchsia',label='North. hemi.')
plt.legend(frameon=False,fontsize=fsize_leg)
ax.set_xlim(5.4,7.2)
ax.set_ylim(0,9.5)
ax.set_xlabel(r'$z$',fontsize=fsize)
ax.set_ylabel('Count',fontsize=fsize)
ax.xaxis.set_minor_locator(AutoMinorLocator())
ax.yaxis.set_minor_locator(AutoMinorLocator())
ax.tick_params(axis='both',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

plt.tight_layout()
plt.savefig('../plots/RLQSO_zdist_north.png')
plt.show()
plt.close()



fig = plt.figure(figsize=(5.,10./1.2))
gs = gridspec.GridSpec(2,1)

ax1 = plt.subplot(gs[0,0])

ax1.hist(z,bins=bins,density=False,histtype='step',linewidth=2,color='royalblue',label='All RLQSO')
ax1.hist(z_north,bins=bins,density=False,histtype='step',linestyle=':',linewidth=2,color='fuchsia',label='North. hemi.')
plt.legend(frameon=False,fontsize=fsize_leg)
ax1.set_xticks(np.arange(5.5,7.1,0.5))
ax1.set_xlim(5.4,7.2)
ax1.set_ylim(0,9.5)
ax1.set_ylabel('Count',fontsize=fsize)
ax1.xaxis.set_minor_locator(AutoMinorLocator())
ax1.yaxis.set_minor_locator(AutoMinorLocator())
ax1.tick_params(axis='x',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=0)
ax1.tick_params(axis='y',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax1.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

ax2 = plt.subplot(gs[1,0],sharex=ax1)

ind_lim = np.where(S147_limit>0)[0]
ind_nolim = np.where(S147_limit<1)[0]

ax2.errorbar(z[ind_lim], S147[ind_lim], yerr=0.2*S147[ind_lim], uplims=True, fmt='o', markersize=7.5, color='royalblue', label='Upper limit')
ax2.errorbar(z[ind_nolim], S147[ind_nolim], uplims=False, fmt='o', markersize=7.5, color='royalblue', label='Detection')
ax2.set_xlim(5.4,7.2)
ax2.set_ylim(0.1,200)
ax2.set_yscale('log')
ax2.set_xlabel(r'$z$',fontsize=fsize)
ax2.set_ylabel(r'$S_{147}\,\rm [mJy]$',fontsize=fsize)
ax2.xaxis.set_minor_locator(AutoMinorLocator())
ax2.tick_params(axis='both',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax2.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

plt.tight_layout()
plt.subplots_adjust(wspace=0, hspace=0)
plt.savefig('../plots/RLQSO_zdist.png')
plt.show()
plt.close()



bins = np.arange(-2.5,70.,5)

fig = plt.figure(figsize=(5.,5.))
gs = gridspec.GridSpec(1,1)

ax = plt.subplot(gs[0,0])

ax.hist(S147,bins=bins,density=False,histtype='step',linewidth=2,color='royalblue',label='All RLQSO')
ax.hist(S147_north,bins=bins,density=False,histtype='step',linestyle=':',linewidth=2,color='fuchsia',label='North. hemi.')
plt.legend(frameon=False,fontsize=fsize_leg)
ax.set_ylim(0,10.5)
ax.set_xlabel(r'$S_{147\,\rm MHz}\,\rm [mJy]$',fontsize=fsize)
ax.set_ylabel('Count',fontsize=fsize)
ax.xaxis.set_minor_locator(AutoMinorLocator())
ax.yaxis.set_minor_locator(AutoMinorLocator())
ax.tick_params(axis='both',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

plt.tight_layout()
plt.savefig('../plots/RLQSO_Sdist_north.png')
plt.show()
plt.close()