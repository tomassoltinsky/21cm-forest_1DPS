'''Functions to incorporate instrumental features

Version 18.10.2023

Spectrum smoothed by running boxcar and fixed boxcar
Gaussian white noise including Aeff/Tsys dependence on freq and intrinsic power-law radio spectrum
'''

import sys
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
import scipy.signal as scisi
from astropy.convolution import convolve, Box1DKernel
from matplotlib.ticker import AutoMinorLocator
from matplotlib import gridspec
import matplotlib.patches as patches
import openpyxl
from numpy import random

#constants
import constants



fsize = 16

def transF(optdep):
  return np.exp(-optdep)

def lambda_obs(redshift,velH):
  return constants.lambda_0*(1+redshift)*(1-velH/constants.c)	#Calculates observed shifted wavelength

def freq_obs(redshift,velH):
  return constants.c/lambda_obs(redshift,velH)

def z_obs(redshift,velH):
  return lambda_obs(redshift,velH)/constants.lambda_0-1.



def smooth_fixedbox(frequency,signal,spec_res,shownpix=False):
   
  Nbins = len(frequency)
  v_pix = (frequency[-1]-frequency[0])/Nbins/1e3
  Npix = int(np.round(spec_res/v_pix,0))
  if shownpix==True:
    print('Convolve over '+str(Npix)+' pixels')

  Nbins_smooth = int(np.floor(Nbins/Npix))
  Nbins = Nbins_smooth*Npix
  ind_smooth = np.arange(0,Nbins-1,Npix)+int(np.floor(Npix/2))
  freq_smooth = frequency[ind_smooth]
  signal_smooth = np.empty(Nbins_smooth)
  for i in range(Nbins_smooth):
    signal_smooth[i] = np.mean(signal[i*Npix:(i+1)*Npix])

  return freq_smooth,signal_smooth

def smooth_runningbox(frequency,signal,spec_res):
   
  Nbins = len(frequency)
  v_pix = (frequency[-1]-frequency[0])/Nbins/1e3
  Npix = int(np.round(spec_res/v_pix,0))
  print('Convolve over '+str(Npix)+' pixels')

  box_kernel = Box1DKernel(Npix)
  signal_smooth = convolve(signal, box_kernel,boundary='fill', fill_value=1.)
  ind_smooth = np.arange(0,Nbins+1,Npix)
  freq_smooth = frequency[ind_smooth]
  signal_smooth = signal_smooth[ind_smooth]

  return freq_smooth,signal_smooth



def excel_column(data_name,n_col):

    wb = openpyxl.load_workbook(data_name)
    sheet = wb.active
    N_rows = sheet.max_row-1

    col = np.empty(N_rows)

    for i_row in range(N_rows):
        col[i_row] = sheet.cell(row=i_row+2,column=n_col).value

    return col

def add_noise(frequency,telescope,dv,S_source,spec_index,t_integration,N_dish,showsigN=False):
  
  #read uGMRT Aeff/Tsys from Fig. 8 in Braun et al. 2019
  datafile = 'sensitivity/sens_%s.xlsx' % telescope
  freq_0 = excel_column(datafile,1)*1000
  ATsys_0 = excel_column(datafile,2)

  if telescope=='uGMRT':
    ATsys_0 = ATsys_0/np.sqrt(30*(30-1))

  elif telescope=='SKA1-low':
    ATsys_0 = ATsys_0/np.sqrt(512*(512-1))

  frequency = frequency/1e6
  index_freq = np.digitize(frequency,freq_0,right=True)	#Interpolate in the density (similar to redshift)

  freq_low  = freq_0[index_freq-1]
  freq_high = freq_0[index_freq]

  w_low = 1./(frequency-freq_low)/(1./(freq_high-frequency)+1./(frequency-freq_low))
  w_low = np.where(frequency==freq_low,1.,w_low)
  w_low = np.where(frequency==freq_high,0.,w_low)

  w_high = 1./(freq_high-frequency)/(1./(freq_high-frequency)+1./(frequency-freq_low))
  w_high = np.where(frequency==freq_high,1.,w_high)
  w_high = np.where(frequency==freq_low,0.,w_high)

  ATsys  = w_low*ATsys_0[index_freq-1]+w_high*ATsys_0[index_freq]

  #calculate standard deviation for telescope following equations 2-3 in Ciardi et al. 2015 MNRAS 453, 101-105; Datta et al. 2007 MNRAS 382, 809–818
  n_noise = 2*constants.k_B/1e7/ATsys/np.sqrt(N_dish*(N_dish-1)*t_integration*3600*dv*1000)/constants.mJy

  #generate radio spectrum and normalize the noise
  S_v = S_source*np.power(frequency/147.,spec_index)
  sigma_noise = n_noise/S_v

  if showsigN==True:

    print('<Noise>       = %.3fmJy' % np.mean(n_noise))
    print('<sigma_noise> = %.5f' % np.mean(sigma_noise))

  #add random values from gaussian distribution to signal
  noise = np.random.normal(0.,sigma_noise,len(frequency))
  return noise



def uni_freq(freq,signal):

  # Calculate frequency range and, because we want to compute the power spectrum 
  # in frequency space, create a uniform frequency array.
  minimum_frequency = np.min(freq)    
  maximum_frequency = np.max(freq)
  Delta_nu = maximum_frequency-minimum_frequency
  dnu = 1000. # Hz
  npix = int((Delta_nu)/dnu)
  uniform_freq = np.linspace(minimum_frequency, maximum_frequency, npix)

  # Initialize flux array and interpolate flux for each line of sight
  N_los = signal.shape[0]
  uniform_signal = np.zeros((N_los, npix), dtype=np.float32)
  for i in range(N_los):
      uniform_signal[i,:] = np.interp(uniform_freq, freq, signal[i])

  return uniform_freq,uniform_signal



#Create N_samples samples of median PS signal from N_obs randomly drawn LOS
def multi_obs(signal,N_obs,N_samples):
   
  N_los,N_kbins = signal.shape
  signal_multiobs = np.empty((N_samples,N_kbins))

  for i in range(N_samples):
    LOS = random.randint(0,N_los-1,size=N_obs)
    signal_multiobs[i][:] = np.mean(signal[LOS][:],axis=0)

  return signal_multiobs